import openpyxl
from secret_store import SecretStore
from config_manager import ConfigManager
from crawler import WebCrawler

# ================= 测试: 授权状态读写 =================


def test_auth_persistence(tmp_path):
    store_path = tmp_path / "test_store.enc"
    key_path = tmp_path / "test_key.key"
    store = SecretStore(key_path=str(key_path), store_path=str(store_path))

    # 默认未授权
    assert store.get_auth_status() is False

    # 写入授权
    store.set_auth_status(True)

    # 重新读取确认持久化
    new_store = SecretStore(key_path=str(key_path), store_path=str(store_path))
    assert new_store.get_auth_status() is True

# ================= 测试: URL 校验 =================


def test_url_validation():
    crawler = WebCrawler(start_url="https://example.com")
    assert crawler.is_valid_url("https://example.com/page1") is True
    assert crawler.is_valid_url("http://test.com") is True
    assert crawler.is_valid_url("javascript:alert(1)") is False
    assert crawler.is_valid_url("mailto:test@test.com") is False

# ================= 测试: skill 匹配算法 =================


def test_skill_matching(tmp_path):
    config_dir = tmp_path / "config"
    skill_dir = config_dir / "skill"
    skill_dir.mkdir(parents=True)

    # 创建一些 YAML 技能文件
    (skill_dir / "请假流程.yaml").write_text('url_keyword: "leave/req"\ndescription: "年假事假申请"', encoding="utf-8")
    (skill_dir / "报销申请.yml").write_text('url_keyword: "expense"\ndescription: "差旅报销"', encoding="utf-8")
    (skill_dir / "重叠测试.yaml").write_text('url_keyword: "leave"\ndescription: "测试"', encoding="utf-8")

    cm = ConfigManager(str(config_dir))

    # 1. 唯一命中
    matched, cands = cm.match_skill("https://oa.com/expense/123")
    assert matched == "报销申请"

    # 2. 多条命中
    matched, cands = cm.match_skill("https://oa.com/leave/req/123")
    assert matched is None
    assert set(cands) == {"请假流程", "重叠测试"}

    # 3. 仍无法唯一确定 (无匹配)
    matched, cands = cm.match_skill("https://oa.com/other")
    assert matched is None
    assert set(cands) == {"请假流程", "报销申请", "重叠测试"}

# ================= 测试: Excel 写入格式 =================


def test_excel_writing(tmp_path):
    from batch_processor import BatchProcessor
    from llm_processor import LLMProcessor

    output_dir = tmp_path / "output"

    # Mock dependencies
    cm = ConfigManager(str(tmp_path))
    llm = LLMProcessor("token", "url", "model")

    bp = BatchProcessor(cm, llm, str(output_dir))

    # 追加一条数据
    bp.append_to_excel("http://test.com", "报销申请", "找到了按钮", "审批通过")

    # 验证 Excel 文件是否按要求写入
    wb = openpyxl.load_workbook(bp.excel_path)
    ws = wb.active

    # 表头
    headers = [cell.value for cell in ws[1]]
    assert headers == ["URL", "匹配技能", "分析证据", "处理结果"]

    # 第一行数据
    row2 = [cell.value for cell in ws[2]]
    assert row2 == ["http://test.com", "报销申请", "找到了按钮", "审批通过"]
